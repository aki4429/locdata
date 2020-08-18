#!/usr/bin/env python
# -*- coding: utf-8 -*-

#縫製品カバー整理場所を設定するために必要なデータを取り込み加工。
#空ラックリスト、maxlist.csv を読みこみ、 到着データchoose_inv.write_c
#とbdataをseiri.py にインプットして、整理リストに加工する。

import csv
import copy
import os
import seiri
import openpyxl

EMPTY = "data/empty_rack.csv"
MAX = "maxlist.csv"
SEXCEL = "arrive_template.xlsx"
#KNFILE = "kizon_new.csv"
#SEIRIFILE = "kizon_new_for_label.csv"

class MakeSeiri:
    def __init__(self, arrivedfile, t_data):
        self.lines = []
        self.af = arrivedfile
        maxlist = self.read_max(MAX)
        empties = self.read_empty(EMPTY)
        self.lines = self.write_seiri(t_data, maxlist, empties)

    def write_seiri(self, t_data, maxlist, empties):
        lines = []
        for row in t_data: 
            code = row[0]
            qty = int(row[1])
            max_q = self.get_max(code, maxlist)
            kizons = []
            i=2
            while(i < len(row)):
                kizons.append([row[i], int(row[i+1])])
                i += 2

            #print(code, qty, kizons)
            s = seiri.Seiri(code, qty, kizons, empties, max_q)
            newlines = s.make_list()
            empties = s.empties
            for newline in newlines:
                lines.append(newline)

        #print("lines:", lines)
        knfile = os.path.join('data', 'KN' + self.af + '.csv')
        with open(knfile, 'w', encoding='CP932') as kn:
            knwriter = csv.writer(kn, lineterminator='\n')
            for line in lines :
                knwriter.writerow([line[0], line[1], line[2]])

        print(knfile, 'を書きました。')

        #カバー整理用に入荷数、既存カートン振り分け数、
        #新規カートン振り分け数を記載したシートを作成します。
        srfile = os.path.join('data', 'SEIRI_' + self.af + '.csv')
        with open(srfile, 'w', encoding='CP932') as sf:
            sfwriter = csv.writer(sf)
            sfwriter.writerows(lines)

        print(srfile, 'を書きました。')

        self.write_excel(lines)

        return lines

    def read_empty(self, filename):
        empties =[]
        with open(filename, 'r', encoding='CP932') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader: 
                empties.append(row[0])

        #print(empties)
        return empties

    def read_max(self, filename):
        maxlist =[]
        with open(filename, 'r') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader: 
                maxlist.append([row[0], int(row[1])])

        #print(maxlist)
        return maxlist

    #コードのピース番号を取り出し、max 数を確定
    def get_max(self, code, maxlist):
        item = code.split("-")[1].split('C')[0]
        #print("item:", item)
        for data in maxlist:
            if data[0] == item:
                return data[1]

    #インボイス入荷カバーを加えて番地リストを保存
    def write_kn(self, bdata, hiduke):
        dc_bdata = copy.deepcopy(bdata)
        for line in self.lines:
            for bd in dc_bdata:
            #ファイルデータとラックデータのアドレスが同じだったら
                if line[2] == bd[0] :
                    #ファイルデータとラックデータのコードも同じだったら
                    if line[0] == bd[1]:
                        bd[2] += int(float(line[1]))
                    elif bd[1] == "empty":
                        bd[1] = line[0]
                        bd[2] = int(float(line[1]))
                    elif line[0] != bd[1]:
                        print(line[2], "のコードが合致しません")
        
        before_filename = os.path.join('stock', 'rev_before_' + self.af +  hiduke.strftime('%Y%m%d') + '.csv')
        after_filename = os.path.join('stock', 'rev_after_' + self.af +  hiduke.strftime('%Y%m%d') + '.csv')

        with open(before_filename, 'w', encoding='CP932') as f:
            writer = csv.writer(f)
            writer.writerows(bdata)
            print(before_filename, 'を保存しました。')

        with open(after_filename, 'w', encoding='CP932') as g:
            writer = csv.writer(g)
            writer.writerows(dc_bdata)
            print(after_filename, 'を保存しました。')

    def write_excel(self, lines):
        wb = openpyxl.load_workbook(SEXCEL)
        sheet = wb['yotei_out']
        sheet['A2'] = "入荷カバーリスト" + self.af

        #コード、数量、番地、数量記入
        n = 0 #5行目からスタート
        while n < len(lines) :
            sheet.cell(row=n+5, column=1, value = lines[n][0]) #code
            sheet.cell(row=n+5, column=2, value = lines[n][1]) #qty
            sheet.cell(row=n+5, column=3, value = lines[n][2]) #banch
            sheet.cell(row=n+5, column=4, value = lines[n][3]) #qty
            n += 1

        filename = os.path.join('data', "Cover_{}.xlsx".format(self.af))
        wb.save(filename)
        print("{}を書きました。".format(filename))





#m = MakeSeiri()
    
