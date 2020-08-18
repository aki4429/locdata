#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import os
import shutil
import datetime as dt
from shiji import shiji
from drop import drop
import sqlite3
import pandas as pd
import openpyxl

DBN = 'locdata.sqlite'
DB = '../odachin.pythonanywhere.com/db.sqlite3'

#棚卸用ラックリスト
RACKLIST = "rack_list.xlsx"
DBEGIN = 4 #データ開始行


#更新日付でcsvファイルを書き出し、データベースも更新
#ファイルが存在するときは、バックアップを保存して上書き
def write_stock_2(hiduke, bdb):
    filename = 'stock/stock_' + hiduke.strftime('%Y%m%d') + '.csv'
    if os.path.isfile(filename):
        ktime = dt.datetime.now().strftime('%y%m%d%s')
        bkfilename = filename + 'bk' + ktime
        shutil.copy(filename, bkfilename)
        print(bkfilename, 'を保存しました。')

    with open(filename, 'w', encoding='CP932') as f:
        writer = csv.writer(f)
        writer.writerows(bdb)

    print(filename, 'を書きました。')

    bdf = pd.DataFrame(bdb)
    bdf.columns =  ['banch', 'code', 'qty'] 

    conn = sqlite3.connect(DBN)
    cur = conn.cursor()

    bdf.to_sql('locdata', conn, if_exists='replace', index=False)
    print(DBN, 'のlocdata テーブルを更新しました。')

    conn.close()

#更新日付でcsvファイルを書き出し、データベースも更新
#ファイルが存在するときは、バックアップを保存して上書き
def write_stock(hiduke, bdb):
    filename = 'stock/stock_' + hiduke.strftime('%Y%m%d') + '.csv'
    if os.path.isfile(filename):
        ktime = dt.datetime.now().strftime('%y%m%d%s')
        bkfilename = filename + 'bk' + ktime
        shutil.copy(filename, bkfilename)
        print(bkfilename, 'を保存しました。')

    with open(filename, 'w', encoding='CP932') as f:
        writer = csv.writer(f)
        writer.writerows(bdb)

    print(filename, 'を書きました。')

    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute('delete from locdata')

    sql = 'insert into locdata (banch, code, qty) values (?,?,?)'
    cur.executemany(sql, bdb)

    conn.commit()
    conn.close()

    print(DB, 'のlocdata テーブルを更新しました。')



#日付をキーにして、コードと使用総数のリストと使用番地指示。
#(参考) dates = key = 日付 + v = コード、製番、数量 のリスト 辞書
#codes =  コード, 数量の辞書をつくつ
def koshin(bdata, days, dates):
    for d in days:
        line = ''
        line += '\n' + d.strftime('%m/%d') +" 計画分 TFC縫製カバー使用番地: "
        line += '\n=================================='
        codes = {} #コード数量のリストを作成
        for v in dates[d]:
            if v[0] not in codes:
                codes[v[0]] = int(v[2])
            else:
                codes[v[0]] += int(v[2])

        sorted_codes = sorted(codes.items())

        for k, val in sorted_codes:
            line += '\n' + k+'('+str(val)+')'+'は、'
            counter = 0
            for shi in shiji(k, int(val), bdata):
                if counter > 0:
                    line += '\n' + ' '*25 
                line += '['+shi[0]+']' +'(' + str(shi[1]) +')から'+ '[' + str(shi[2]) + ']を使用してください。'  
                counter += 1

        line += '\n-------------------------------------'
        line += '\nコード (数量)  [製番]'
        for data in dates[d]:
            line += '\n'+ data[0] + '(' + data[2] + ') ' + '[' + data[1] +']'

        print(line)
        drop(bdata)

        write_stock(d, bdata)

    return bdata

def write_racklist(bdata, koshinbi):
    book = openpyxl.load_workbook(RACKLIST)
    sheet = book['racklist']
    k_bi = koshinbi.strftime('%y%m%d')
    sheet["B1"].value = "{}更新のラックリスト".format(k_bi)
    i = DBEGIN
    for row in bdata:
        sheet.cell(row = i, column = 1).value = row[0]
        sheet.cell(row = i, column = 2).value = row[1]
        sheet.cell(row = i, column = 3).value = row[2]
        i+=1

    filename = "racklilst_{}.xlsx".format(k_bi)
    book.save(filename)
    print("{}を書きました。".format(filename))




